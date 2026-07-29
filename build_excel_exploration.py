"""
Builds the Phase 1 Excel exploration workbook: raw data sample, a
pivot-style monthly summary built with live formulas, and a log of
data quality issues found during the initial look at the data.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import os

# Set this to your name -- otherwise Excel will list "openpyxl" as the author
AUTHOR_NAME = "Chaitra"

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
orders = pd.read_csv(os.path.join(BASE, "data", "orders.csv"))
items = pd.read_csv(os.path.join(BASE, "data", "order_items.csv"))

# Join to get one row per order-item with status + month
orders_small = orders[["order_id", "order_status", "order_purchase_ts"]].copy()
merged = items.merge(orders_small, on="order_id", how="left")
merged["order_month"] = pd.to_datetime(merged["order_purchase_ts"], errors="coerce").dt.strftime("%Y-%m")
merged["line_total"] = merged["unit_price"] * merged["quantity"] + merged["freight_value"]

wb = openpyxl.Workbook()

HEADER_FILL = PatternFill(start_color="4C72B0", end_color="4C72B0", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)

def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

# ---------------------------------------------------------------
# Sheet 1: Raw data sample (first 500 rows, so file stays light)
# ---------------------------------------------------------------
ws1 = wb.active
ws1.title = "raw_order_items"
sample = merged[["order_id", "product_id", "quantity", "unit_price",
                  "freight_value", "line_total", "order_status", "order_month"]].head(500)

ws1.append(list(sample.columns))
style_header(ws1, len(sample.columns))
for row in sample.itertuples(index=False):
    ws1.append(list(row))
for c in range(1, len(sample.columns) + 1):
    ws1.column_dimensions[get_column_letter(c)].width = 16

last_row = len(sample) + 1  # +1 for header

# ---------------------------------------------------------------
# Sheet 2: Pivot-style summary (built with SUMIFS/COUNTIFS formulas
# referencing the raw sheet -- recalculates if raw data changes)
# ---------------------------------------------------------------
ws2 = wb.create_sheet("monthly_summary")
months = sorted(sample["order_month"].dropna().unique())

ws2.append(["order_month", "num_line_items", "total_revenue", "avg_order_value"])
style_header(ws2, 4)

for i, m in enumerate(months, start=2):
    ws2.cell(row=i, column=1, value=m)
    ws2.cell(row=i, column=2,
        value=f'=COUNTIFS(raw_order_items!H2:H{last_row},A{i})')
    ws2.cell(row=i, column=3,
        value=f'=ROUND(SUMIFS(raw_order_items!F2:F{last_row},raw_order_items!H2:H{last_row},A{i}),2)')
    ws2.cell(row=i, column=4,
        value=f'=ROUND(C{i}/B{i},2)')

for c in range(1, 5):
    ws2.column_dimensions[get_column_letter(c)].width = 18

# ---------------------------------------------------------------
# Sheet 3: Data quality notes from the initial review
# ---------------------------------------------------------------
ws3 = wb.create_sheet("data_quality_notes")
notes = [
    ["Issue", "Table", "Detail", "Fix applied"],
    ["Duplicate rows", "customers", "15 duplicate customer_id rows found", "Dropped in load_data.py with drop_duplicates()"],
    ["Missing values", "customers", "20 rows missing customer_city", "Left as NULL; excluded from city-level breakdowns"],
    ["Missing timestamps", "orders", "25 orders missing order_purchase_ts", "Dropped from time-based analysis (can't bucket by month)"],
    ["Canceled orders", "orders", "~5% of orders have status=canceled", "Excluded from revenue queries (WHERE order_status='delivered')"],
    ["No delivery data", "orders", "Non-delivered orders have null delivery_days", "Excluded from delivery-time analysis"],
]
for r in notes:
    ws3.append(r)
style_header(ws3, 4)
for c in range(1, 5):
    ws3.column_dimensions[get_column_letter(c)].width = 22
for row in ws3.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

for ws in [ws1, ws2, ws3]:
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.font.bold is not True:
                cell.font = BODY_FONT

wb.properties.creator = AUTHOR_NAME
wb.properties.lastModifiedBy = AUTHOR_NAME

out_path = os.path.join(BASE, "outputs", "phase1_excel_exploration.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")
